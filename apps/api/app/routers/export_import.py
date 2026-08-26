"""Export and import classes and question sets as JSON/CSV/Excel."""
import io
import re
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, field_validator
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

# Die Standard-Notenskala stand hier ein zweites Mal (wortgleich, nur ohne
# Unterstrich im Namen) und in noten.py ein drittes Mal mit Text-Schluesseln.
# Es gibt eine: die in scoring.py, wo auch gerechnet wird.
from ..scoring import DEFAULT_SCALE, bewerte, status_of
from ..schueler import sortiert
from ..pdfdruck import als_anhang, neue_seite
from ..austauschformat import quiz_inhalt, quiz_schnappschuss
from ..felder import ohne_leer, ohne_none
from ..besitz import oder_403
from ..database import get_db
from ..importe import geprueft
from ..models import SchoolClass, Student, QuestionSet, QuestionSetItem, Question, Session, Scan, Folder, User
from .auth import get_current_user, rate_limit


def strip_latex(text: str) -> str:
    """Convert simple LaTeX to readable plain text for PDFs."""
    s = re.sub(r'\$([^$]+)\$', r'\1', text)
    s = s.replace('\\frac{', '(').replace('}{', ')/(').replace('}', ')')
    for cmd, repl in [('\\cdot', '·'), ('\\pm', '±'), ('\\neq', '≠'),
                       ('\\leq', '≤'), ('\\geq', '≥'), ('\\pi', 'π'),
                       ('\\infty', '∞'), ('\\sqrt', '√'), ('\\sum', 'Σ')]:
        s = s.replace(cmd, repl)
    s = re.sub(r'\^{([^}]*)}', r'^\1', s)
    s = re.sub(r'_{([^}]*)}', r'_\1', s)
    s = re.sub(r'\\[a-zA-Z]+\s*', '', s)
    return s.strip()

from .modules import modul_pflicht
from .folders import ensure_set_access
from .noten import _grade_from_pct as _dezimalnote

# Klassen-Export/-Import gehoeren dem Kern, Fragen/Sets/Sitzungen dem Modul.
# Deshalb haengt die Schranke hier an der einzelnen Route, nicht am Router.
CARDVOTE = Depends(modul_pflicht("cardvote"))

router = APIRouter(prefix="/api", tags=["export"])


async def _quiz_flags(db, session):
    """E/G-Differenzierung und Minuspunkte des Quiz hinter einer Session."""
    if not session.question_set_id:
        return False, False
    qs = await db.get(QuestionSet, session.question_set_id)
    if not qs:
        return False, False
    return bool(qs.niveau_aktiv), bool(qs.minuspunkte)


async def _session_items(db, session):
    """Die Set-Eintraege einer Session in Reihenfolge — eine Abfrage, vier Stellen."""
    if not session.question_set_id:
        return []
    result = await db.execute(
        select(QuestionSetItem)
        .options(selectinload(QuestionSetItem.question))
        .where(QuestionSetItem.question_set_id == session.question_set_id)
        .order_by(QuestionSetItem.position)
    )
    return list(result.scalars().all())


async def _session_questions(db, session):
    """Fragen einer Session als ORM-Objekte, mit gemischter Loesung und Niveau.

    Stand wortgleich in evaluation_xlsx und evaluation_scsv. Die beiden
    Zusatzfelder haengen bewusst am Objekt (nicht in der DB): sie gelten nur
    fuer diese Session, weil question_map die Antworten gemischt hat."""
    questions = []
    qmap = session.question_map or {}
    for item in await _session_items(db, session):
        q = item.question
        q._shuffled_correct = qmap.get(str(q.id), q.correct_answer)
        q._niveau = item.niveau or ""
        questions.append(q)
    return questions


async def _session_question_dicts(db, session):
    """Fragen einer Session als Dicts mit Text — fuer die PDFs.

    Stand wortgleich in student_evaluation_pdf und all_students_pdf. Die
    aehnlich aussehende Fassung in results.py hat kein "text" (andere
    Nutzlast) und bleibt bewusst getrennt."""
    qmap = session.question_map or {}
    return [
        {"id": item.question.id, "text": item.question.text,
         "correct_answer": qmap.get(str(item.question.id), item.question.correct_answer),
         "niveau": item.niveau or ""}
        for item in await _session_items(db, session)
    ]


def _question_from_import(qdata, owner_id) -> Question:
    """Frage aus einer Importdatei. Stand wortgleich in import_question_set
    und _import_folder_recursive."""
    return Question(
        text=qdata.text,
        choices=qdata.choices,
        correct_answer=qdata.correct_answer,
        image_url=qdata.image_url,
        image_layout=qdata.image_layout,
        num_choices=qdata.num_choices,
        choice_images=qdata.choice_images,
        owner_id=owner_id,
    )


XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _xlsx_response(wb, filename: str) -> StreamingResponse:
    """Arbeitsmappe als Download. Stand dreimal wortgleich, nur der Dateiname
    war verschieden."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


def _xlsx_template(title: str, headers: list, ausrichten: bool = False):
    """Leere Vorlage mit formatierter Kopfzeile (Workbook + ws).

    Stand zweimal fast wortgleich; die Klassenvorlage richtet ihre Kopfzeile
    zusaetzlich aus, die Fragenvorlage nicht — deshalb der Schalter, damit
    beide Dateien Zelle fuer Zelle bleiben, was sie waren."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = title

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(fgColor="F5F5F7", fill_type="solid")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        if ausrichten:
            cell.alignment = Alignment(horizontal="center" if col == 1 else "left")
    return wb, ws


# --- Export ---

@router.get("/export/class/{class_id}")
async def export_class(class_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(SchoolClass).options(selectinload(SchoolClass.students)).where(SchoolClass.id == class_id)
    )
    cls = result.scalar_one_or_none()
    if not cls:
        raise HTTPException(404)
    if cls.owner_id and cls.owner_id != user.id:
        raise HTTPException(403, "Kein Zugriff auf diese Klasse")
    return {
        "type": "cardvote_class",
        "version": 1,
        "name": cls.name,
        # Bewusst nur card_id und name: niveau, foerder und notizen sind
        # besonders schuetzenswert (DSGVO Art. 9) und gehoeren in keine Datei,
        # die weitergegeben wird. Wer sie hier ergaenzt, macht aus einem
        # Klassenexport eine Foerderakte.
        "students": [{"card_id": s.card_id, "name": s.name} for s in sorted(cls.students, key=lambda s: s.card_id)],
    }


@router.get("/export/question-set/{set_id}", dependencies=[CARDVOTE])
async def export_question_set(set_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    qs = await db.get(QuestionSet, set_id)
    if not qs:
        raise HTTPException(404)
    await ensure_set_access(db, qs, user.id)
    result = await db.execute(
        select(QuestionSetItem)
        .options(selectinload(QuestionSetItem.question))
        .where(QuestionSetItem.question_set_id == set_id)
        .order_by(QuestionSetItem.position)
    )
    items = result.scalars().all()
    return quiz_schnappschuss(qs, items)


# --- Excel template for class import ---

@router.get("/import/class-template.xlsx")
async def class_xlsx_template():
    from openpyxl.styles import Alignment

    wb, ws = _xlsx_template("Klasse", ["Karten-Nr", "Name"], ausrichten=True)

    for i in range(1, 36):
        ws.cell(row=i + 1, column=1, value=i).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30

    return _xlsx_response(wb, "CardVote_Vorlage.xlsx")


# --- Excel Import for classes ---

MAX_XLSX_BYTES = 5 * 1024 * 1024


def _arbeitsblatt(data: bytes):
    """Excel-Datei oeffnen und das erste Blatt liefern.

    Ohne diesen Mantel endete jede Datei, die keine .xlsx ist (eine .csv, eine
    abgeschnittene Uebertragung, eine .numbers-Datei) als HTTP 500 mit einem
    openpyxl-Traceback — die Lehrkraft erfuhr nicht, dass schlicht das Format
    nicht passt."""
    from openpyxl import load_workbook
    try:
        wb = load_workbook(io.BytesIO(data))
    except Exception:
        raise HTTPException(400, "Die Datei laesst sich nicht lesen — bitte eine Excel-Datei (.xlsx) hochladen.")
    ws = wb.active
    if ws is None:
        raise HTTPException(400, "Die Excel-Datei enthaelt kein Tabellenblatt")
    return ws


def _skala(config: dict) -> dict:
    """Notenskala aus der Session-Konfiguration, robust gelesen.

    Ueberall stand ``{int(k): v for k, v in scale_raw.items()}`` — eine
    Konfiguration mit einem unlesbaren Schluessel (Altbestand, von Hand
    bearbeitet) liess damit den ganzen Export mit HTTP 500 auffliegen."""
    roh = (config or {}).get("grade_scale") or DEFAULT_SCALE
    try:
        skala = {int(k): float(v) for k, v in roh.items()}
    except (AttributeError, TypeError, ValueError):
        return dict(DEFAULT_SCALE)
    return skala if all(g in skala for g in (1, 2, 3, 4, 5, 6)) else dict(DEFAULT_SCALE)


@router.post("/import/class-xlsx")
async def import_class_xlsx(name: str = "Neue Klasse", file: UploadFile = File(...), user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rate_limit("import", f"u{user.id}", 60, 3600, "Zu viele Importe. Bitte kurz warten.")
    data = await file.read(MAX_XLSX_BYTES + 1)
    if len(data) > MAX_XLSX_BYTES:
        raise HTTPException(400, "Datei zu gross (max 5 MB)")
    ws = _arbeitsblatt(data)

    sc = SchoolClass(name=name, owner_id=user.id)
    db.add(sc)
    await db.flush()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        card_id_val, name_val = row[0], row[1]
        if not name_val or not str(name_val).strip():
            continue
        try:
            card_id = int(card_id_val)
        except (ValueError, TypeError):
            continue
        db.add(Student(card_id=card_id, name=str(name_val).strip(), class_id=sc.id))
        count += 1

    if count == 0:
        raise HTTPException(400, "Keine Lernenden in der Excel-Datei gefunden")

    await db.commit()
    await db.refresh(sc)
    return {"id": sc.id, "name": sc.name, "count": count}


# --- Excel template for question set import ---

@router.get("/import/questions-template.xlsx", dependencies=[CARDVOTE])
async def questions_xlsx_template():
    headers = ["Frage", "Antwort A", "Antwort B", "Antwort C", "Antwort D", "Richtig (z.B. A oder AB)"]
    wb, ws = _xlsx_template("Fragen", headers)

    ws.cell(row=2, column=1, value="Was ist 2+2?")
    ws.cell(row=2, column=2, value="3")
    ws.cell(row=2, column=3, value="4")
    ws.cell(row=2, column=4, value="5")
    ws.cell(row=2, column=5, value="6")
    ws.cell(row=2, column=6, value="B")

    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["F"].width = 22

    return _xlsx_response(wb, "CardVote_Fragen_Vorlage.xlsx")


# --- Excel Import for question sets ---

@router.post("/import/questions-xlsx", dependencies=[CARDVOTE])
async def import_questions_xlsx(name: str = "Neues Frageset", folder_id: Optional[int] = None, file: UploadFile = File(...), user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rate_limit("import", f"u{user.id}", 60, 3600, "Zu viele Importe. Bitte kurz warten.")
    data = await file.read(MAX_XLSX_BYTES + 1)
    if len(data) > MAX_XLSX_BYTES:
        raise HTTPException(400, "Datei zu gross (max 5 MB)")
    ws = _arbeitsblatt(data)

    # owner_id: sonst ist das importierte Set fuer JEDES Konto lesbar
    # (ensure_set_access laesst owner-lose Sets als Altbestand durch).
    qs = QuestionSet(name=name, folder_id=folder_id, owner_id=user.id)
    db.add(qs)
    await db.flush()

    pos = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not str(row[0]).strip():
            continue
        text = str(row[0]).strip()
        choices = {
            "A": str(row[1]).strip() if len(row) > 1 and row[1] else "",
            "B": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            "C": str(row[3]).strip() if len(row) > 3 and row[3] else "",
            "D": str(row[4]).strip() if len(row) > 4 and row[4] else "",
        }
        correct = str(row[5]).strip().upper() if len(row) > 5 and row[5] else ""
        num_choices = 4
        if not choices["D"] and not choices["C"]:
            num_choices = 2
        elif not choices["D"]:
            num_choices = 3

        q = Question(text=text, choices=choices, correct_answer=correct, num_choices=num_choices, owner_id=user.id)
        db.add(q)
        await db.flush()
        db.add(QuestionSetItem(question_set_id=qs.id, question_id=q.id, position=pos))
        pos += 1

    if pos == 0:
        raise HTTPException(400, "Keine Fragen in der Excel-Datei gefunden")

    await db.commit()
    return {"id": qs.id, "name": qs.name, "count": pos}


# --- JSON Import ---

class ImportStudent(BaseModel):
    """Ein Kind aus der Klassendatei. Frueher ``int(s["card_id"])`` roh — eine
    Datei ohne Karten-Nummer endete als HTTP 500 mit Traceback."""
    # card_id und name waren immer Pflicht — ohne sie warf der Import einen
    # KeyError (HTTP 500). Jetzt sagt die Meldung, welches Feld fehlt.
    card_id: int
    name: str

    @field_validator("name", mode="before")
    @classmethod
    def _text(cls, v):
        return "" if v is None else str(v)


class ImportClassBody(BaseModel):
    type: str = ""
    name: str = ""
    students: List[ImportStudent] = []

    _leer_text = field_validator("name", mode="before")(ohne_none(""))
    _leer_liste = field_validator("students", mode="before")(ohne_none([]))


class ImportQuestion(BaseModel):
    """Eine Frage aus der Datei. ``text`` fehlte -> KeyError -> HTTP 500."""
    text: str
    choices: dict = {"A": "", "B": "", "C": "", "D": ""}
    correct_answer: Optional[str] = None
    image_url: Optional[str] = None
    image_layout: str = "above"
    num_choices: int = 4
    choice_images: Optional[dict] = None
    niveau: str = ""

    _leer_text = field_validator("text", "image_layout", "niveau", mode="before")(ohne_none(""))
    _leer_choices = field_validator("choices", mode="before")(ohne_none({"A": "", "B": "", "C": "", "D": ""}))
    _leer_zahl = field_validator("num_choices", mode="before")(ohne_leer(4))

    @field_validator("image_layout")
    @classmethod
    def _layout(cls, v):
        return v or "above"


class ImportQuestionSetBody(BaseModel):
    type: str = ""
    name: str = ""
    folder_id: Optional[int] = None
    shuffle_questions: bool = False
    shuffle_answers: bool = False
    # E/G und Minuspunkte gehoeren zum Quiz — ohne sie waere die importierte
    # Fassung anders bewertet als das Original.
    niveau_aktiv: bool = False
    minuspunkte: bool = False
    questions: List[ImportQuestion] = []

    _leer_text = field_validator("name", mode="before")(ohne_none(""))
    _leer_flag = field_validator("shuffle_questions", "shuffle_answers", "niveau_aktiv",
                                 "minuspunkte", mode="before")(ohne_none(False))
    _leer_liste = field_validator("questions", mode="before")(ohne_none([]))


@router.post("/import/class")
async def import_class(body: dict, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """`body: dict` in der Signatur ist Absicht — geprueft() antwortet mit 400
    und Feldnamen statt FastAPIs englischer 422 (siehe app/importe.py)."""
    rate_limit("import", f"u{user.id}", 60, 3600, "Zu viele Importe. Bitte kurz warten.")
    if not isinstance(body, dict) or body.get("type") != "cardvote_class":
        raise HTTPException(400, "Ungültiges Format")
    if isinstance(body.get("students"), list) and len(body["students"]) > 50:
        raise HTTPException(400, "Maximal 50 Lernende pro Klasse")
    daten = geprueft(ImportClassBody, body, "Klassendatei")
    sc = SchoolClass(name=daten.name[:200], owner_id=user.id)
    db.add(sc)
    await db.flush()
    for s in daten.students:
        name = s.name.strip()[:200]
        if s.card_id < 0 or s.card_id > 49 or not name:
            continue
        db.add(Student(card_id=s.card_id, name=name, class_id=sc.id))
    await db.commit()
    await db.refresh(sc)
    return {"id": sc.id, "name": sc.name}


@router.post("/import/question-set", dependencies=[CARDVOTE])
async def import_question_set(body: dict, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """`body: dict` in der Signatur ist Absicht — siehe app/importe.py."""
    rate_limit("import", f"u{user.id}", 60, 3600, "Zu viele Importe. Bitte kurz warten.")
    if not isinstance(body, dict) or body.get("type") != "cardvote_questionset":
        raise HTTPException(400, "Ungültiges Format")
    if isinstance(body.get("questions"), list) and len(body["questions"]) > 200:
        raise HTTPException(400, "Maximal 200 Fragen pro Set")
    daten = geprueft(ImportQuestionSetBody, body, "Fragendatei")
    qs = QuestionSet(
        name=daten.name,
        folder_id=daten.folder_id,
        owner_id=user.id,   # sonst fuer jedes Konto lesbar (Altbestand-Ausnahme)
        shuffle_questions=daten.shuffle_questions,
        shuffle_answers=daten.shuffle_answers,
        niveau_aktiv=daten.niveau_aktiv,
        minuspunkte=daten.minuspunkte,
    )
    db.add(qs)
    await db.flush()
    for pos, qdata in enumerate(daten.questions):
        q = _question_from_import(qdata, user.id)
        db.add(q)
        await db.flush()
        db.add(QuestionSetItem(question_set_id=qs.id, question_id=q.id, position=pos,
                               niveau="E" if qdata.niveau == "E" else ""))
    await db.commit()
    return {"id": qs.id, "name": qs.name}


# --- Folder export/import ---

async def _export_folder_recursive(folder_id: int, db: AsyncSession) -> dict:
    folder = await db.get(Folder, folder_id)
    if not folder:
        raise HTTPException(404)
    result = await db.execute(
        select(QuestionSet).where(QuestionSet.folder_id == folder_id)
    )
    sets = result.scalars().all()
    exported_sets = []
    for qs in sets:
        items_r = await db.execute(
            select(QuestionSetItem)
            .options(selectinload(QuestionSetItem.question))
            .where(QuestionSetItem.question_set_id == qs.id)
            .order_by(QuestionSetItem.position)
        )
        items = items_r.scalars().all()
        # Ohne "type"/"version": die stehen einmal aussen am Ordner, nicht an
        # jedem Quiz darin. Der Inhalt selbst ist derselbe wie beim Einzelexport.
        exported_sets.append(quiz_inhalt(qs, items))
    children_r = await db.execute(
        select(Folder).where(Folder.parent_id == folder_id)
    )
    children = children_r.scalars().all()
    exported_children = []
    for child in children:
        exported_children.append(await _export_folder_recursive(child.id, db))
    return {
        "name": folder.name,
        "question_sets": exported_sets,
        "children": exported_children,
    }


@router.get("/export/folder/{folder_id}", dependencies=[CARDVOTE])
async def export_folder(folder_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    folder = await db.get(Folder, folder_id)
    if not folder:
        raise HTTPException(404)
    if folder.owner_id and folder.owner_id != user.id:
        raise HTTPException(403, "Kein Zugriff auf diesen Ordner")
    data = await _export_folder_recursive(folder_id, db)
    return {"type": "cardvote_folder", "version": 1, **data}


class ImportFolderSet(BaseModel):
    name: str = ""
    shuffle_questions: bool = False
    shuffle_answers: bool = False
    niveau_aktiv: bool = False
    minuspunkte: bool = False
    questions: List[ImportQuestion] = []

    _leer_text = field_validator("name", mode="before")(ohne_none(""))
    _leer_flag = field_validator("shuffle_questions", "shuffle_answers", "niveau_aktiv",
                                 "minuspunkte", mode="before")(ohne_none(False))
    _leer_liste = field_validator("questions", mode="before")(ohne_none([]))


class ImportFolderBody(BaseModel):
    type: str = ""
    version: int = 1
    name: str = "Ordner"
    question_sets: List[ImportFolderSet] = []
    children: List["ImportFolderBody"] = []

    _leer_text = field_validator("name", mode="before")(ohne_none(""))
    _leer_liste = field_validator("question_sets", "children", mode="before")(ohne_none([]))

    @field_validator("name")
    @classmethod
    def _name_ok(cls, v):
        return v.strip()[:200] or "Ordner"


ImportFolderBody.model_rebuild()


async def _import_folder_recursive(data: ImportFolderBody, parent_id, owner_id, db: AsyncSession, ziel: Folder = None):
    """Ordner der Datei anlegen — oder den Inhalt in einen vorhandenen legen.

    `ziel` ist die Antwort der Lehrkraft auf „neu anlegen oder mit Vorhandenem
    verknuepfen?": ist es gesetzt, entsteht KEIN zweiter gleichnamiger Ordner,
    die Sets und Unterordner der Datei landen direkt darin. Nur die oberste
    Ebene kennt das; darunter wird immer angelegt, sonst wuerde ein Import
    fremde Unterordner zusammenlegen, die nur zufaellig gleich heissen.
    """
    if ziel is not None:
        folder = ziel
    else:
        folder = Folder(name=data.name, parent_id=parent_id, owner_id=owner_id)
        db.add(folder)
        await db.flush()
    for qs_data in data.question_sets:
        qs = QuestionSet(
            name=qs_data.name,
            folder_id=folder.id,
            # owner_id: sonst fuer jedes Konto lesbar. Hier stand frueher
            # `user.id` — in dieser Funktion gibt es kein `user`, jeder
            # Ordner-Import endete daher als NameError, also HTTP 500.
            owner_id=owner_id,
            shuffle_questions=qs_data.shuffle_questions,
            shuffle_answers=qs_data.shuffle_answers,
            niveau_aktiv=qs_data.niveau_aktiv,
            minuspunkte=qs_data.minuspunkte,
        )
        db.add(qs)
        await db.flush()
        for pos, qdata in enumerate(qs_data.questions):
            q = _question_from_import(qdata, owner_id)
            db.add(q)
            await db.flush()
            db.add(QuestionSetItem(question_set_id=qs.id, question_id=q.id, position=pos,
                                   niveau="E" if qdata.niveau == "E" else ""))
    for child_data in data.children:
        await _import_folder_recursive(child_data, folder.id, owner_id, db)
    return folder


@router.post("/import/folder", dependencies=[CARDVOTE])
async def import_folder(body: dict, folder_id: Optional[int] = None, in_folder: bool = False,
                        db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    """`body: dict` in der Signatur ist Absicht — siehe app/importe.py."""
    rate_limit("import", f"u{user.id}", 60, 3600, "Zu viele Importe. Bitte kurz warten.")
    if not isinstance(body, dict) or body.get("type") != "cardvote_folder":
        raise HTTPException(400, "Ungültiges Format")

    def _count(node):
        if not isinstance(node, dict):
            return 0
        sets = node.get("question_sets") or []
        sets = sets if isinstance(sets, list) else []
        n = 0
        for s in sets:
            fragen = (s.get("questions") if isinstance(s, dict) else None) or []
            n += len(fragen) if isinstance(fragen, list) else 0
        kinder = node.get("children") or []
        for child in (kinder if isinstance(kinder, list) else []):
            n += _count(child)
        return n
    if _count(body) > 5000:
        raise HTTPException(400, "Import zu gross (max. 5000 Fragen pro Ordner)")

    ziel = None
    if folder_id is not None:
        # Ohne diese Pruefung landete der Import im Ordner eines fremden Kontos
        # — oder in keinem, und der Fremdschluessel warf einen HTTP 500.
        ziel = await db.get(Folder, folder_id)
        if not ziel or (ziel.owner_id and ziel.owner_id != user.id):
            raise HTTPException(404, "Ordner nicht gefunden")

    daten = geprueft(ImportFolderBody, body, "Ordnerdatei")
    # in_folder: der Inhalt kommt in den gewaehlten Ordner selbst, statt darin
    # einen zweiten gleichnamigen anzulegen.
    folder = await _import_folder_recursive(daten, folder_id, user.id, db,
                                            ziel=ziel if (in_folder and ziel is not None) else None)
    await db.commit()
    return {"id": folder.id, "name": folder.name}


# --- Duplicate question set ---

@router.post("/question-sets/{set_id}/duplicate", dependencies=[CARDVOTE])
async def duplicate_question_set(set_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(QuestionSet)
        .options(selectinload(QuestionSet.items).selectinload(QuestionSetItem.question))
        .where(QuestionSet.id == set_id)
    )
    orig = result.scalar_one_or_none()
    if not orig:
        raise HTTPException(404)
    await ensure_set_access(db, orig, user.id)

    qs = QuestionSet(
        name=f"{orig.name} (Kopie)",
        folder_id=orig.folder_id,
        owner_id=user.id,   # die Kopie gehoert der Person, die sie anlegt
        shuffle_questions=orig.shuffle_questions,
        shuffle_answers=orig.shuffle_answers,
    )
    db.add(qs)
    await db.flush()

    for item in orig.items:
        db.add(QuestionSetItem(question_set_id=qs.id, question_id=item.question_id, position=item.position))

    await db.commit()
    return {"id": qs.id, "name": qs.name}


# --- Excel export for evaluation ---

@router.get("/sessions/{session_id}/evaluation-xlsx", dependencies=[CARDVOTE])
async def evaluation_xlsx(session_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    session = await oder_403(db, Session, session_id, user,
                        verboten="Kein Zugriff auf diese Session")

    students = []
    if session.class_id:
        # Klassenreihenfolge wie auf dem Bildschirm: position. card_id ist die
        # Nummer der gedruckten ArUco-Karte, keine Reihenfolge (app/schueler.py).
        students = await sortiert(db, Student.class_id == session.class_id)

    questions = await _session_questions(db, session)

    result = await db.execute(select(Scan).where(Scan.session_id == session_id))
    all_scans = result.scalars().all()
    scan_map = {(s.student_id, s.question_id): s.answer for s in all_scans}

    wb = Workbook()
    ws = wb.active
    ws.title = "Auswertung"

    header_font = Font(bold=True, size=11)
    green_fill = PatternFill(fgColor="D4EDDA", fill_type="solid")
    red_fill = PatternFill(fgColor="FDE2D9", fill_type="solid")

    headers = ["Name"] + [f"F{i+1}" for i in range(len(questions))] + ["Punkte", "%"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Solution row
    ws.cell(row=2, column=1, value="Lösung").font = Font(bold=True, color="888888")
    for i, q in enumerate(questions):
        ws.cell(row=2, column=i + 2, value=q._shuffled_correct or "–")

    config = session.eval_config or {}
    niveau_aktiv, minuspunkte = await _quiz_flags(db, session)
    qdicts = [{"id": q.id, "correct_answer": q._shuffled_correct, "niveau": getattr(q, "_niveau", "")} for q in questions]
    row = 3
    for student in students:
        has_any = any((student.card_id, q.id) in scan_map for q in questions)
        # Wer als krank gilt, steht nicht in der Liste; eine gewertete 0 schon.
        if status_of(student.card_id, has_any, config) == "krank":
            continue
        ws.cell(row=row, column=1, value=student.name).font = Font(bold=True)
        for i, q in enumerate(questions):
            answer = scan_map.get((student.card_id, q.id))
            cell = ws.cell(row=row, column=i + 2, value=answer or "–")
            correct = q._shuffled_correct
            if answer and correct:
                if answer in correct:
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
            cell.alignment = Alignment(horizontal="center")
        eigene = {q.id: scan_map.get((student.card_id, q.id)) for q in questions}
        w = bewerte(qdicts, eigene, niveau=student.niveau or "", niveau_aktiv=niveau_aktiv,
                    minuspunkte=minuspunkte, weights=config.get("weights"), scale=config.get("grade_scale"))
        ws.cell(row=row, column=len(questions) + 2, value=f"{w['score']:g}/{w['max_score']:g}")
        ws.cell(row=row, column=len(questions) + 3, value=f"{round(w['pct'])}%")
        row += 1

    ws.column_dimensions["A"].width = 20

    return _xlsx_response(wb, f"CardVote_Auswertung_{session_id}.xlsx")


# --- iDoceo SCSV export ---

@router.get("/sessions/{session_id}/evaluation-scsv", dependencies=[CARDVOTE])
async def evaluation_scsv(session_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Export as semicolon-separated CSV for iDoceo import."""
    session = await oder_403(db, Session, session_id, user,
                        verboten="Kein Zugriff auf diese Session")

    students = []
    if session.class_id:
        # Klassenreihenfolge wie auf dem Bildschirm: position. card_id ist die
        # Nummer der gedruckten ArUco-Karte, keine Reihenfolge (app/schueler.py).
        students = await sortiert(db, Student.class_id == session.class_id)

    questions = await _session_questions(db, session)

    result = await db.execute(select(Scan).where(Scan.session_id == session_id))
    all_scans = result.scalars().all()
    scan_map = {(s.student_id, s.question_id): s.answer for s in all_scans}

    config = session.eval_config or {}
    niveau_aktiv, minuspunkte = await _quiz_flags(db, session)
    weights = config.get("weights", {})
    scale = _skala(config)
    # Gewichte gehen unten direkt an bewerte(weights=...); kein eigener Zugriff nötig.

    esc = lambda v: f'"{v}"'

    set_name = ""
    if session.question_set_id:
        qs_obj = await db.get(QuestionSet, session.question_set_id)
        if qs_obj:
            set_name = qs_obj.name

    last_scan_result = await db.execute(
        select(Scan.scanned_at).where(Scan.session_id == session_id).order_by(Scan.scanned_at.desc()).limit(1)
    )
    last_scan_date = last_scan_result.scalar_one_or_none()
    test_date = last_scan_date or session.created_at
    date_str = test_date.strftime("%d.%m.%Y") if test_date else ""
    title = f"{set_name} ({date_str})" if set_name else f"Test {session_id} ({date_str})"

    lines = []
    header = [esc(""), esc(title), esc(""), esc("")]
    lines.append(",".join(header))

    scanned_question_ids = set(qid for (_, qid) in scan_map)

    qdicts = [{"id": qn.id, "correct_answer": qn._shuffled_correct, "niveau": getattr(qn, "_niveau", "")}
              for qn in questions if qn.id in scanned_question_ids]
    for student in students:
        has_any = any((student.card_id, qn.id) in scan_map for qn in questions)
        if status_of(student.card_id, has_any, config) == "krank":
            continue
        eigene = {qn["id"]: scan_map.get((student.card_id, qn["id"])) for qn in qdicts}
        pct = round(bewerte(qdicts, eigene, niveau=student.niveau or "", niveau_aktiv=niveau_aktiv,
                            minuspunkte=minuspunkte, weights=weights, scale=scale)["pct"])
        grade = _decimal_grade(pct, scale)
        lines.append(",".join([esc(student.name), esc(str(grade)), esc(""), esc("")]))

    content = "\n".join(lines)
    buf = io.BytesIO(content.encode("utf-8-sig"))
    filename = f"CardVote_{session_id}.csv"
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- Individual student evaluation PDFs ---

def _grade_from_pct(pct, scale=None):
    s = scale or DEFAULT_SCALE
    for g in range(1, 6):
        if pct >= s.get(g, s.get(str(g), 0)):
            return g
    return 6


def _decimal_grade(pct, scale=None):
    """Prozent -> Dezimalnote. Eine Quelle: die Fassung aus noten.py.

    Hier stand eine zweite Umrechnung, die mit Pythons round() rundete — und
    das rundet die halbe Stelle zur GERADEN Zahl ("Bankers Rounding"). 83,5 %
    ergaben im Export 2,2, im Notenbuch und auf dem Bildschirm aber 2,3.
    Dieselbe Leistung stand also im PDF als andere Note als in der Anwendung,
    und zwar bei jeder Prozentzahl, die genau auf einer halben Stelle liegt.

    Gerechnet wird jetzt an einer Stelle; kaufmaennisch runden steht in
    scoring.py und ist dort begruendet."""
    return _dezimalnote(pct, scale or DEFAULT_SCALE)


def _antwort_zeilen(questions, card_id, scan_map, get_w):
    """Antworten eines Kindes je Frage: Zeilen fuer die Tabelle und die
    Abgaben fuer bewerte(). Stand wortgleich im Einzel- und im Sammel-PDF."""
    results = []
    eigene = {}
    for q in questions:
        ans = scan_map.get((card_id, q["id"]))
        eigene[q["id"]] = ans
        correct = q["correct_answer"]
        is_correct = ans and correct and ans in correct
        w = get_w(q["id"])
        results.append({"text": q["text"], "answer": ans, "correct": correct, "is_correct": is_correct, "weight": w})
    return results, eigene


def _antwort_farbe(c, r):
    """Zeilenfarbe im PDF: richtig gruen, falsch rot, keine Abgabe grau.
    Stand wortgleich im Einzel- und im Sammel-PDF."""
    from reportlab.lib.colors import HexColor
    if r["is_correct"]:
        c.setFillColor(HexColor("#0a7d3e"))
    elif r["answer"] and r["correct"]:
        c.setFillColor(HexColor("#d1350f"))
    else:
        c.setFillColorRGB(0.4, 0.4, 0.4)


def _build_student_pdf_single(student, questions, scan_map, session, config, niveau_aktiv=False, minuspunkte=False):
    from reportlab.lib.units import mm
    # HexColor faellt hier weg: die Farbwahl steckt jetzt in _antwort_farbe().

    buf = io.BytesIO()
    # A4-Leinwand aus app/pdfdruck.py — dieselben Zeilen standen an acht Stellen.
    c, pw, ph = neue_seite(buf)

    weights = config.get("weights", {}) if config else {}
    scale = _skala(config or {})
    # eval_config["times"] (Dauer je Frage) zeigt nur die Auswertungsseite;
    # das PDF nennt allein die Gesamtdauer.
    total_time = config.get("total_time") if config else None

    get_w = lambda qid: weights.get(str(qid), weights.get(qid, 1))
    # max_score kommt weiter unten aus bewerte() — eine Quelle für die Wertung.

    results, eigene = _antwort_zeilen(questions, student["card_id"], scan_map, get_w)

    # Punkte, Prozent und damit die Note kommen aus der gemeinsamen Wertung
    # (E/G-Bonus, Minuspunkte) — sonst stuende im PDF etwas anderes als am Schirm.
    wertung = bewerte(questions, eigene, niveau=student.get("niveau", ""),
                      niveau_aktiv=niveau_aktiv, minuspunkte=minuspunkte,
                      weights=weights, scale=scale)
    score, max_score, pct = wertung["score"], wertung["max_score"], round(wertung["pct"])
    grade = _decimal_grade(pct, scale)

    y = ph - 30 * mm
    c.setFont("Helvetica-Bold", 18)
    c.drawString(20 * mm, y, f"Auswertung — {student['name']}")
    y -= 8 * mm
    c.setFont("Helvetica", 11)
    c.drawString(20 * mm, y, session.name or f"Session #{session.id}")
    if total_time:
        c.drawRightString(pw - 20 * mm, y, f"Dauer: {total_time // 60}:{total_time % 60:02d}")
    y -= 12 * mm

    c.setFont("Helvetica-Bold", 14)
    c.drawString(20 * mm, y, f"Note: {grade}    —    {score}/{max_score} Punkte ({pct}%)")
    y -= 14 * mm

    # Table header
    c.setFont("Helvetica-Bold", 9)
    cols = [20 * mm, 35 * mm, pw - 80 * mm, pw - 55 * mm, pw - 35 * mm]
    c.drawString(cols[0], y, "#")
    c.drawString(cols[1], y, "Frage")
    c.drawString(cols[2], y, "Antwort")
    c.drawString(cols[3], y, "Lösung")
    c.drawString(cols[4], y, "Punkte")
    y -= 2 * mm
    c.setStrokeColorRGB(0, 0, 0)
    c.line(20 * mm, y, pw - 20 * mm, y)
    y -= 5 * mm

    c.setFont("Helvetica", 9)
    for i, r in enumerate(results):
        if y < 25 * mm:
            c.showPage()
            y = ph - 25 * mm
            c.setFont("Helvetica", 9)

        _antwort_farbe(c, r)

        text = strip_latex(r["text"])
        text = text[:60] + ("…" if len(text) > 60 else "")
        c.drawString(cols[0], y, str(i + 1))
        c.drawString(cols[1], y, text)
        c.drawString(cols[2], y, r["answer"] or "–")
        c.drawString(cols[3], y, r["correct"] or "–")
        pts = r["weight"] if r["is_correct"] else 0
        c.drawString(cols[4], y, str(pts))
        c.setFillColorRGB(0, 0, 0)
        y -= 5 * mm

    c.save()
    buf.seek(0)
    return buf


@router.get("/sessions/{session_id}/student-pdf/{card_id}", dependencies=[CARDVOTE])
async def student_evaluation_pdf(session_id: int, card_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    session = await oder_403(db, Session, session_id, user,
                        verboten="Kein Zugriff auf diese Session")

    student = None
    if session.class_id:
        result = await db.execute(
            select(Student).where(Student.class_id == session.class_id, Student.card_id == card_id)
        )
        s = result.scalar_one_or_none()
        if s:
            student = {"card_id": s.card_id, "name": s.name, "niveau": s.niveau or ""}
    if not student:
        raise HTTPException(404, "Lernende/r nicht gefunden")

    questions = await _session_question_dicts(db, session)

    result = await db.execute(select(Scan).where(Scan.session_id == session_id))
    scan_map = {(s.student_id, s.question_id): s.answer for s in result.scalars().all()}

    config = session.eval_config or {}
    questions = [q for q in questions if (student["card_id"], q["id"]) in scan_map]
    niveau_aktiv, minuspunkte = await _quiz_flags(db, session)
    buf = _build_student_pdf_single(student, questions, scan_map, session, config, niveau_aktiv, minuspunkte)
    filename = f"Auswertung_{student['name']}_{session_id}.pdf"
    return als_anhang(buf, filename)


@router.get("/sessions/{session_id}/all-students-pdf", dependencies=[CARDVOTE])
async def all_students_pdf(session_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    session = await oder_403(db, Session, session_id, user,
                        verboten="Kein Zugriff auf diese Session")

    students = []
    if session.class_id:
        # Klassenreihenfolge wie auf dem Bildschirm: position (app/schueler.py).
        students = [{"card_id": s.card_id, "name": s.name, "niveau": s.niveau or ""}
                    for s in await sortiert(db, Student.class_id == session.class_id)]

    questions = await _session_question_dicts(db, session)

    result = await db.execute(select(Scan).where(Scan.session_id == session_id))
    all_scans = result.scalars().all()
    scan_map = {(s.student_id, s.question_id): s.answer for s in all_scans}

    from reportlab.lib.units import mm
    # HexColor faellt hier weg: die Farbwahl steckt jetzt in _antwort_farbe().

    config = session.eval_config or {}
    niveau_aktiv, minuspunkte = await _quiz_flags(db, session)
    weights = config.get("weights", {})
    scale = _skala(config)
    get_w = lambda qid: weights.get(str(qid), weights.get(qid, 1))
    # Die erreichbare Punktzahl steht je Kind in wertung["max_score"] (student_max) —
    # mit E/G-Bonus kann sie sich unterscheiden, ein Gesamtwert wäre hier falsch.

    # Wer als krank gilt, kommt nicht ins Sammel-PDF; eine gewertete 0 schon.
    present = [s for s in students
               if status_of(s["card_id"], any((s["card_id"], q["id"]) in scan_map for q in questions), config) != "krank"]

    buf = io.BytesIO()
    # A4-Leinwand aus app/pdfdruck.py — dieselben Zeilen standen an acht Stellen.
    c, pw, ph = neue_seite(buf)

    for si, student in enumerate(present):
        y = ph - 30 * mm
        c.setFont("Helvetica-Bold", 16)
        c.drawString(20 * mm, y, student["name"])
        y -= 7 * mm
        c.setFont("Helvetica", 10)
        c.drawString(20 * mm, y, session.name or f"Session #{session.id}")
        y -= 10 * mm

        student_questions = [q for q in questions if (student["card_id"], q["id"]) in scan_map]
        results, eigene = _antwort_zeilen(student_questions, student["card_id"], scan_map, get_w)

        wertung = bewerte(student_questions, eigene, niveau=student.get("niveau", ""),
                          niveau_aktiv=niveau_aktiv, minuspunkte=minuspunkte,
                          weights=weights, scale=scale)
        score, student_max, pct = wertung["score"], wertung["max_score"], round(wertung["pct"])
        grade = _decimal_grade(pct, scale)

        c.setFont("Helvetica-Bold", 13)
        c.drawString(20 * mm, y, f"Note: {grade}    —    {score}/{student_max} Punkte ({pct}%)")
        y -= 12 * mm

        c.setFont("Helvetica-Bold", 9)
        c.drawString(20 * mm, y, "#")
        c.drawString(30 * mm, y, "Frage")
        c.drawString(pw - 65 * mm, y, "Antw.")
        c.drawString(pw - 45 * mm, y, "Lösung")
        c.drawString(pw - 28 * mm, y, "Pkt.")
        y -= 2 * mm
        c.line(20 * mm, y, pw - 20 * mm, y)
        y -= 5 * mm

        c.setFont("Helvetica", 9)
        for i, r in enumerate(results):
            _antwort_farbe(c, r)

            text = strip_latex(r["text"])
            text = text[:55] + ("…" if len(text) > 55 else "")
            c.drawString(20 * mm, y, str(i + 1))
            c.drawString(30 * mm, y, text)
            c.drawString(pw - 65 * mm, y, r["answer"] or "–")
            c.drawString(pw - 45 * mm, y, r["correct"] or "–")
            pts = r["weight"] if r["is_correct"] else 0
            c.drawString(pw - 28 * mm, y, str(pts))
            c.setFillColorRGB(0, 0, 0)
            y -= 5 * mm

        if si < len(present) - 1:
            c.showPage()

    c.save()
    buf.seek(0)
    filename = f"CardVote_Auswertungen_{session_id}.pdf"
    return als_anhang(buf, filename)


@router.get("/classes/{class_id}/all-tests-student-pdf/{card_id}", dependencies=[CARDVOTE])
async def class_student_pdf(class_id: int, card_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    school_class = await db.get(SchoolClass, class_id)
    if not school_class:
        raise HTTPException(404)
    if school_class.owner_id and school_class.owner_id != user.id:
        raise HTTPException(403, "Kein Zugriff auf diese Klasse")

    result = await db.execute(select(Student).where(Student.class_id == class_id, Student.card_id == card_id))
    student_obj = result.scalar_one_or_none()
    if not student_obj:
        raise HTTPException(404, "Lernende/r nicht gefunden")

    result = await db.execute(select(Session).where(Session.class_id == class_id).order_by(Session.created_at))
    sessions = result.scalars().all()

    from reportlab.lib.units import mm
    from reportlab.lib.colors import HexColor

    buf = io.BytesIO()
    # A4-Leinwand aus app/pdfdruck.py — dieselben Zeilen standen an acht Stellen.
    c, pw, ph = neue_seite(buf)

    y = ph - 30 * mm
    c.setFont("Helvetica-Bold", 18)
    c.drawString(20 * mm, y, f"Gesamtübersicht — {student_obj.name}")
    y -= 7 * mm
    c.setFont("Helvetica", 11)
    c.drawString(20 * mm, y, f"Klasse {school_class.name}")
    y -= 14 * mm

    # Table header
    c.setFont("Helvetica-Bold", 10)
    c.drawString(20 * mm, y, "Test")
    c.drawString(pw - 85 * mm, y, "Punkte")
    c.drawString(pw - 55 * mm, y, "%")
    c.drawString(pw - 35 * mm, y, "Note")
    y -= 2 * mm
    c.line(20 * mm, y, pw - 20 * mm, y)
    y -= 6 * mm

    total_score = 0
    total_max = 0
    test_count = 0

    c.setFont("Helvetica", 10)
    for session in sessions:
        # Nur die Abfrage ist dieselbe wie in den PDFs (_session_items); die
        # Uebersicht braucht keinen Fragetext, deshalb hier eigene Dicts.
        qmap = session.question_map or {}
        questions = [{"id": item.question.id,
                      "correct_answer": qmap.get(str(item.question.id), item.question.correct_answer),
                      "niveau": item.niveau or ""}
                     for item in await _session_items(db, session)]

        scan_result = await db.execute(select(Scan).where(Scan.session_id == session.id, Scan.student_id == card_id))
        scans = {s.question_id: s.answer for s in scan_result.scalars().all()}

        if not scans:
            continue

        config = session.eval_config or {}
        weights = config.get("weights", {})
        scale = _skala(config)
        get_w = lambda qid: weights.get(str(qid), weights.get(qid, 1))

        max_sc = sum(get_w(q["id"]) for q in questions if q["correct_answer"])
        sc = 0
        for q in questions:
            ans = scans.get(q["id"])
            if ans and q["correct_answer"] and ans in q["correct_answer"]:
                sc += get_w(q["id"])

        pct = round(sc / max_sc * 100) if max_sc > 0 else 0
        grade = _decimal_grade(pct, scale)

        total_score += sc
        total_max += max_sc
        test_count += 1

        # Get set name
        set_name = session.name
        if session.question_set_id:
            qs = await db.get(QuestionSet, session.question_set_id)
            if qs:
                set_name = qs.name

        if y < 25 * mm:
            c.showPage()
            y = ph - 25 * mm
            c.setFont("Helvetica", 10)

        label = set_name[:45] + ("…" if len(set_name) > 45 else "")
        c.drawString(20 * mm, y, label)
        c.drawString(pw - 85 * mm, y, f"{sc}/{max_sc}")
        c.drawString(pw - 55 * mm, y, f"{pct}%")

        g_color = "#0a7d3e" if grade <= 2 else "#b8860b" if grade <= 4 else "#d1350f"
        c.setFillColor(HexColor(g_color))
        c.drawString(pw - 35 * mm, y, str(grade))
        c.setFillColorRGB(0, 0, 0)
        y -= 6 * mm

    # Summary
    y -= 4 * mm
    c.line(20 * mm, y, pw - 20 * mm, y)
    y -= 8 * mm
    total_pct = round(total_score / total_max * 100) if total_max > 0 else 0
    total_grade = _decimal_grade(total_pct)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, y, f"Gesamt: {total_score}/{total_max} ({total_pct}%)    Note: {total_grade}    ({test_count} Tests)")

    c.save()
    buf.seek(0)
    filename = f"CardVote_{student_obj.name}_Gesamtübersicht.pdf"
    return als_anhang(buf, filename)
