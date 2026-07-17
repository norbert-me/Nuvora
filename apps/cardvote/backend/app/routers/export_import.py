"""Export and import classes and question sets as JSON/CSV/Excel."""
import io
import re
from typing import Optional

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse, Response
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..database import get_db
from ..models import SchoolClass, Student, QuestionSet, QuestionSetItem, Question, Session, Scan, Folder, User
from .auth import get_current_user, rate_limit


def strip_latex(text: str) -> str:
    """Convert simple LaTeX to readable plain text for PDFs."""
    s = re.sub(r'\$([^$]+)\$', r'\1', text)
    s = s.replace('\\frac{', '(').replace('}{', ')/(').replace('}', ')')
    for cmd, repl in [('\\cdot', '·'), ('\\pm', '±'), ('\\neq', '≠'),
                       ('\\leq', '≤'), ('\\geq', '≥'), ('\\pi', 'π'),
                       ('\\infty', '∞'), ('\\sqrt', '√'), ('\\sum', 'Σ')]:
        s = s.replace(cmd, repl)
    s = re.sub(r'\^{([^}]*)}', r'^\1', s)
    s = re.sub(r'_{([^}]*)}', r'_\1', s)
    s = re.sub(r'\\[a-zA-Z]+\s*', '', s)
    return s.strip()

router = APIRouter(prefix="/api", tags=["export"])


# --- Export ---

@router.get("/export/class/{class_id}")
async def export_class(class_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(SchoolClass).options(selectinload(SchoolClass.students)).where(SchoolClass.id == class_id)
    )
    cls = result.scalar_one_or_none()
    if not cls:
        raise HTTPException(404)
    if cls.owner_id and cls.owner_id != user.id:
        raise HTTPException(403, "Kein Zugriff auf diese Klasse")
    return {
        "type": "cardvote_class",
        "version": 1,
        "name": cls.name,
        "students": [{"card_id": s.card_id, "name": s.name} for s in sorted(cls.students, key=lambda s: s.card_id)],
    }


@router.get("/export/question-set/{set_id}")
async def export_question_set(set_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    qs = await db.get(QuestionSet, set_id)
    if not qs:
        raise HTTPException(404)
    from .folders import ensure_set_access
    await ensure_set_access(db, qs, user.id)
    result = await db.execute(
        select(QuestionSetItem)
        .options(selectinload(QuestionSetItem.question))
        .where(QuestionSetItem.question_set_id == set_id)
        .order_by(QuestionSetItem.position)
    )
    items = result.scalars().all()
    return {
        "type": "cardvote_questionset",
        "version": 1,
        "name": qs.name,
        "shuffle_questions": qs.shuffle_questions,
        "shuffle_answers": qs.shuffle_answers,
        "questions": [
            {
                "text": item.question.text,
                "choices": item.question.choices,
                "correct_answer": item.question.correct_answer,
                "image_url": item.question.image_url,
                "image_layout": item.question.image_layout,
                "num_choices": item.question.num_choices,
                "choice_images": item.question.choice_images,
            }
            for item in items
        ],
    }


# --- Excel template for class import ---

@router.get("/import/class-template.xlsx")
async def class_xlsx_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Klasse"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(fgColor="F5F5F7", fill_type="solid")
    for col, h in enumerate(["Karten-Nr", "Name"], 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center" if col == 1 else "left")

    for i in range(1, 36):
        ws.cell(row=i + 1, column=1, value=i).alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="CardVote_Vorlage.xlsx"'},
    )


# --- Excel Import for classes ---

MAX_XLSX_BYTES = 5 * 1024 * 1024


@router.post("/import/class-xlsx")
async def import_class_xlsx(name: str = "Neue Klasse", file: UploadFile = File(...), user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rate_limit("import", f"u{user.id}", 60, 3600, "Zu viele Importe. Bitte kurz warten.")
    from openpyxl import load_workbook

    data = await file.read(MAX_XLSX_BYTES + 1)
    if len(data) > MAX_XLSX_BYTES:
        raise HTTPException(400, "Datei zu gross (max 5 MB)")
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    sc = SchoolClass(name=name, owner_id=user.id)
    db.add(sc)
    await db.flush()

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 2:
            continue
        card_id_val, name_val = row[0], row[1]
        if not name_val or not str(name_val).strip():
            continue
        try:
            card_id = int(card_id_val)
        except (ValueError, TypeError):
            continue
        db.add(Student(card_id=card_id, name=str(name_val).strip(), class_id=sc.id))
        count += 1

    if count == 0:
        raise HTTPException(400, "Keine Lernenden in der Excel-Datei gefunden")

    await db.commit()
    await db.refresh(sc)
    return {"id": sc.id, "name": sc.name, "count": count}


# --- Excel template for question set import ---

@router.get("/import/questions-template.xlsx")
async def questions_xlsx_template():
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Fragen"

    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(fgColor="F5F5F7", fill_type="solid")
    headers = ["Frage", "Antwort A", "Antwort B", "Antwort C", "Antwort D", "Richtig (z.B. A oder AB)"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    ws.cell(row=2, column=1, value="Was ist 2+2?")
    ws.cell(row=2, column=2, value="3")
    ws.cell(row=2, column=3, value="4")
    ws.cell(row=2, column=4, value="5")
    ws.cell(row=2, column=5, value="6")
    ws.cell(row=2, column=6, value="B")

    ws.column_dimensions["A"].width = 30
    for col in ["B", "C", "D", "E"]:
        ws.column_dimensions[col].width = 18
    ws.column_dimensions["F"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="CardVote_Fragen_Vorlage.xlsx"'},
    )


# --- Excel Import for question sets ---

@router.post("/import/questions-xlsx")
async def import_questions_xlsx(name: str = "Neues Frageset", folder_id: Optional[int] = None, file: UploadFile = File(...), user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rate_limit("import", f"u{user.id}", 60, 3600, "Zu viele Importe. Bitte kurz warten.")
    from openpyxl import load_workbook

    data = await file.read(MAX_XLSX_BYTES + 1)
    if len(data) > MAX_XLSX_BYTES:
        raise HTTPException(400, "Datei zu gross (max 5 MB)")
    wb = load_workbook(io.BytesIO(data))
    ws = wb.active

    qs = QuestionSet(name=name, folder_id=folder_id)
    db.add(qs)
    await db.flush()

    pos = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0] or not str(row[0]).strip():
            continue
        text = str(row[0]).strip()
        choices = {
            "A": str(row[1]).strip() if len(row) > 1 and row[1] else "",
            "B": str(row[2]).strip() if len(row) > 2 and row[2] else "",
            "C": str(row[3]).strip() if len(row) > 3 and row[3] else "",
            "D": str(row[4]).strip() if len(row) > 4 and row[4] else "",
        }
        correct = str(row[5]).strip().upper() if len(row) > 5 and row[5] else ""
        num_choices = 4
        if not choices["D"] and not choices["C"]:
            num_choices = 2
        elif not choices["D"]:
            num_choices = 3

        q = Question(text=text, choices=choices, correct_answer=correct, num_choices=num_choices, owner_id=user.id)
        db.add(q)
        await db.flush()
        db.add(QuestionSetItem(question_set_id=qs.id, question_id=q.id, position=pos))
        pos += 1

    if pos == 0:
        raise HTTPException(400, "Keine Fragen in der Excel-Datei gefunden")

    await db.commit()
    return {"id": qs.id, "name": qs.name, "count": pos}


# --- JSON Import ---

class ImportClassBody(BaseModel):
    type: str
    name: str
    students: list


class ImportQuestionSetBody(BaseModel):
    type: str
    name: str
    folder_id: Optional[int] = None
    shuffle_questions: bool = False
    shuffle_answers: bool = False
    questions: list


@router.post("/import/class")
async def import_class(body: ImportClassBody, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rate_limit("import", f"u{user.id}", 60, 3600, "Zu viele Importe. Bitte kurz warten.")
    if body.type != "cardvote_class":
        raise HTTPException(400, "Ungültiges Format")
    if len(body.students) > 50:
        raise HTTPException(400, "Maximal 50 Lernende pro Klasse")
    sc = SchoolClass(name=body.name[:200], owner_id=user.id)
    db.add(sc)
    await db.flush()
    for s in body.students:
        card_id = int(s["card_id"])
        name = str(s["name"]).strip()[:200]
        if card_id < 0 or card_id > 49 or not name:
            continue
        db.add(Student(card_id=card_id, name=name, class_id=sc.id))
    await db.commit()
    await db.refresh(sc)
    return {"id": sc.id, "name": sc.name}


@router.post("/import/question-set")
async def import_question_set(body: ImportQuestionSetBody, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    rate_limit("import", f"u{user.id}", 60, 3600, "Zu viele Importe. Bitte kurz warten.")
    if body.type != "cardvote_questionset":
        raise HTTPException(400, "Ungültiges Format")
    if len(body.questions) > 200:
        raise HTTPException(400, "Maximal 200 Fragen pro Set")
    qs = QuestionSet(
        name=body.name,
        folder_id=body.folder_id,
        shuffle_questions=body.shuffle_questions,
        shuffle_answers=body.shuffle_answers,
    )
    db.add(qs)
    await db.flush()
    for pos, qdata in enumerate(body.questions):
        q = Question(
            text=qdata["text"],
            choices=qdata.get("choices", {"A": "", "B": "", "C": "", "D": ""}),
            correct_answer=qdata.get("correct_answer"),
            image_url=qdata.get("image_url"),
            image_layout=qdata.get("image_layout", "above"),
            num_choices=qdata.get("num_choices", 4),
            choice_images=qdata.get("choice_images"),
            owner_id=user.id,
        )
        db.add(q)
        await db.flush()
        db.add(QuestionSetItem(question_set_id=qs.id, question_id=q.id, position=pos))
    await db.commit()
    return {"id": qs.id, "name": qs.name}


# --- Folder export/import ---

async def _export_folder_recursive(folder_id: int, db: AsyncSession) -> dict:
    folder = await db.get(Folder, folder_id)
    if not folder:
        raise HTTPException(404)
    result = await db.execute(
        select(QuestionSet).where(QuestionSet.folder_id == folder_id)
    )
    sets = result.scalars().all()
    exported_sets = []
    for qs in sets:
        items_r = await db.execute(
            select(QuestionSetItem)
            .options(selectinload(QuestionSetItem.question))
            .where(QuestionSetItem.question_set_id == qs.id)
            .order_by(QuestionSetItem.position)
        )
        items = items_r.scalars().all()
        exported_sets.append({
            "name": qs.name,
            "shuffle_questions": qs.shuffle_questions,
            "shuffle_answers": qs.shuffle_answers,
            "questions": [
                {
                    "text": item.question.text,
                    "choices": item.question.choices,
                    "correct_answer": item.question.correct_answer,
                    "image_url": item.question.image_url,
                    "image_layout": item.question.image_layout,
                    "num_choices": item.question.num_choices,
                    "choice_images": item.question.choice_images,
                }
                for item in items
            ],
        })
    children_r = await db.execute(
        select(Folder).where(Folder.parent_id == folder_id)
    )
    children = children_r.scalars().all()
    exported_children = []
    for child in children:
        exported_children.append(await _export_folder_recursive(child.id, db))
    return {
        "name": folder.name,
        "question_sets": exported_sets,
        "children": exported_children,
    }


@router.get("/export/folder/{folder_id}")
async def export_folder(folder_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    folder = await db.get(Folder, folder_id)
    if not folder:
        raise HTTPException(404)
    if folder.owner_id and folder.owner_id != user.id:
        raise HTTPException(403, "Kein Zugriff auf diesen Ordner")
    data = await _export_folder_recursive(folder_id, db)
    return {"type": "cardvote_folder", "version": 1, **data}


async def _import_folder_recursive(data: dict, parent_id, owner_id, db: AsyncSession):
    folder = Folder(name=data["name"], parent_id=parent_id, owner_id=owner_id)
    db.add(folder)
    await db.flush()
    for qs_data in data.get("question_sets", []):
        qs = QuestionSet(
            name=qs_data["name"],
            folder_id=folder.id,
            shuffle_questions=qs_data.get("shuffle_questions", False),
            shuffle_answers=qs_data.get("shuffle_answers", False),
        )
        db.add(qs)
        await db.flush()
        for pos, qdata in enumerate(qs_data.get("questions", [])):
            q = Question(
                text=qdata["text"],
                choices=qdata.get("choices", {"A": "", "B": "", "C": "", "D": ""}),
                correct_answer=qdata.get("correct_answer"),
                image_url=qdata.get("image_url"),
                image_layout=qdata.get("image_layout", "above"),
                num_choices=qdata.get("num_choices", 4),
                choice_images=qdata.get("choice_images"),
                owner_id=owner_id,
            )
            db.add(q)
            await db.flush()
            db.add(QuestionSetItem(question_set_id=qs.id, question_id=q.id, position=pos))
    for child_data in data.get("children", []):
        await _import_folder_recursive(child_data, folder.id, owner_id, db)
    return folder


@router.post("/import/folder")
async def import_folder(body: dict, folder_id: Optional[int] = None, db: AsyncSession = Depends(get_db), user: User = Depends(get_current_user)):
    rate_limit("import", f"u{user.id}", 60, 3600, "Zu viele Importe. Bitte kurz warten.")
    if body.get("type") != "cardvote_folder":
        raise HTTPException(400, "Ungültiges Format")

    def _count(node):
        sets = node.get("question_sets", []) or []
        n = sum(len(s.get("questions", []) or []) for s in sets)
        for child in (node.get("children", []) or []):
            n += _count(child)
        return n
    if _count(body) > 5000:
        raise HTTPException(400, "Import zu gross (max. 5000 Fragen pro Ordner)")

    folder = await _import_folder_recursive(body, folder_id, user.id, db)
    await db.commit()
    return {"id": folder.id, "name": folder.name}


# --- Duplicate question set ---

@router.post("/question-sets/{set_id}/duplicate")
async def duplicate_question_set(set_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(QuestionSet)
        .options(selectinload(QuestionSet.items).selectinload(QuestionSetItem.question))
        .where(QuestionSet.id == set_id)
    )
    orig = result.scalar_one_or_none()
    if not orig:
        raise HTTPException(404)
    from .folders import ensure_set_access
    await ensure_set_access(db, orig, user.id)

    qs = QuestionSet(
        name=f"{orig.name} (Kopie)",
        folder_id=orig.folder_id,
        shuffle_questions=orig.shuffle_questions,
        shuffle_answers=orig.shuffle_answers,
    )
    db.add(qs)
    await db.flush()

    for item in orig.items:
        db.add(QuestionSetItem(question_set_id=qs.id, question_id=item.question_id, position=item.position))

    await db.commit()
    return {"id": qs.id, "name": qs.name}


# --- Excel export for evaluation ---

@router.get("/sessions/{session_id}/evaluation-xlsx")
async def evaluation_xlsx(session_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    session = await db.get(Session, session_id)
    if session and session.owner_id and session.owner_id != user.id:
        raise HTTPException(403, "Kein Zugriff auf diese Session")
    if not session:
        raise HTTPException(404)

    students = []
    if session.class_id:
        result = await db.execute(
            select(Student).where(Student.class_id == session.class_id).order_by(Student.name)
        )
        students = result.scalars().all()

    questions = []
    qmap = session.question_map or {}
    if session.question_set_id:
        result = await db.execute(
            select(QuestionSetItem)
            .options(selectinload(QuestionSetItem.question))
            .where(QuestionSetItem.question_set_id == session.question_set_id)
            .order_by(QuestionSetItem.position)
        )
        for item in result.scalars().all():
            q = item.question
            q._shuffled_correct = qmap.get(str(q.id), q.correct_answer)
            questions.append(q)

    result = await db.execute(select(Scan).where(Scan.session_id == session_id))
    all_scans = result.scalars().all()
    scan_map = {(s.student_id, s.question_id): s.answer for s in all_scans}

    wb = Workbook()
    ws = wb.active
    ws.title = "Auswertung"

    header_font = Font(bold=True, size=11)
    green_fill = PatternFill(fgColor="D4EDDA", fill_type="solid")
    red_fill = PatternFill(fgColor="FDE2D9", fill_type="solid")

    headers = ["Name"] + [f"F{i+1}" for i in range(len(questions))] + ["Punkte", "%"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    # Solution row
    ws.cell(row=2, column=1, value="Lösung").font = Font(bold=True, color="888888")
    for i, q in enumerate(questions):
        ws.cell(row=2, column=i + 2, value=q._shuffled_correct or "–")

    row = 3
    for student in students:
        has_any = any((student.card_id, q.id) in scan_map for q in questions)
        if not has_any:
            continue
        ws.cell(row=row, column=1, value=student.name).font = Font(bold=True)
        score = 0
        total_scored = 0
        for i, q in enumerate(questions):
            answer = scan_map.get((student.card_id, q.id))
            cell = ws.cell(row=row, column=i + 2, value=answer or "–")
            correct = q._shuffled_correct
            if answer and correct:
                total_scored += 1
                if answer in correct:
                    score += 1
                    cell.fill = green_fill
                else:
                    cell.fill = red_fill
            cell.alignment = Alignment(horizontal="center")
        max_score = len([q for q in questions if q._shuffled_correct])
        ws.cell(row=row, column=len(questions) + 2, value=f"{score}/{max_score}")
        ws.cell(row=row, column=len(questions) + 3, value=f"{round(score/max_score*100)}%" if max_score > 0 else "–")
        row += 1

    ws.column_dimensions["A"].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"CardVote_Auswertung_{session_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- iDoceo SCSV export ---

@router.get("/sessions/{session_id}/evaluation-scsv")
async def evaluation_scsv(session_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    """Export as semicolon-separated CSV for iDoceo import."""
    session = await db.get(Session, session_id)
    if not session:
        raise HTTPException(404)
    if session.owner_id and session.owner_id != user.id:
        raise HTTPException(403, "Kein Zugriff auf diese Session")

    students = []
    if session.class_id:
        result = await db.execute(
            select(Student).where(Student.class_id == session.class_id).order_by(Student.name)
        )
        students = result.scalars().all()

    questions = []
    qmap = session.question_map or {}
    if session.question_set_id:
        result = await db.execute(
            select(QuestionSetItem)
            .options(selectinload(QuestionSetItem.question))
            .where(QuestionSetItem.question_set_id == session.question_set_id)
            .order_by(QuestionSetItem.position)
        )
        for item in result.scalars().all():
            q = item.question
            q._shuffled_correct = qmap.get(str(q.id), q.correct_answer)
            questions.append(q)

    result = await db.execute(select(Scan).where(Scan.session_id == session_id))
    all_scans = result.scalars().all()
    scan_map = {(s.student_id, s.question_id): s.answer for s in all_scans}

    config = session.eval_config or {}
    weights = config.get("weights", {})
    get_w = lambda qid: weights.get(str(qid), weights.get(qid, 1))
    scale_raw = config.get("grade_scale", {1: 87, 2: 73, 3: 59, 4: 45, 5: 20, 6: 0})
    scale = {int(k): v for k, v in scale_raw.items()}

    esc = lambda v: f'"{v}"'

    set_name = ""
    if session.question_set_id:
        qs_obj = await db.get(QuestionSet, session.question_set_id)
        if qs_obj:
            set_name = qs_obj.name

    last_scan_result = await db.execute(
        select(Scan.scanned_at).where(Scan.session_id == session_id).order_by(Scan.scanned_at.desc()).limit(1)
    )
    last_scan_date = last_scan_result.scalar_one_or_none()
    test_date = last_scan_date or session.created_at
    date_str = test_date.strftime("%d.%m.%Y") if test_date else ""
    title = f"{set_name} ({date_str})" if set_name else f"Test {session_id} ({date_str})"

    lines = []
    header = [esc(""), esc(title), esc(""), esc("")]
    lines.append(",".join(header))

    scanned_question_ids = set(qid for (_, qid) in scan_map)

    for student in students:
        has_any = any((student.card_id, qn.id) in scan_map for qn in questions)
        if not has_any:
            continue
        student_max = sum(
            get_w(qn.id) for qn in questions
            if qn._shuffled_correct and qn.id in scanned_question_ids
        )
        score = sum(
            get_w(qn.id) for qn in questions
            if qn._shuffled_correct and scan_map.get((student.card_id, qn.id))
            and scan_map[(student.card_id, qn.id)] in qn._shuffled_correct
        )
        pct = round(score / student_max * 100) if student_max > 0 else 0
        grade = _decimal_grade(pct, scale)
        lines.append(",".join([esc(student.name), esc(str(grade)), esc(""), esc("")]))

    content = "\n".join(lines)
    buf = io.BytesIO(content.encode("utf-8-sig"))
    filename = f"CardVote_{session_id}.csv"
    return StreamingResponse(
        buf,
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# --- Individual student evaluation PDFs ---

DEFAULT_SCALE = {1: 87, 2: 73, 3: 59, 4: 45, 5: 20, 6: 0}


def _grade_from_pct(pct, scale=None):
    s = scale or DEFAULT_SCALE
    for g in range(1, 6):
        if pct >= s.get(g, s.get(str(g), 0)):
            return g
    return 6


def _decimal_grade(pct, scale=None):
    s = scale or DEFAULT_SCALE
    ranges = [
        (1, s.get(1, s.get("1", 87)), 100),
        (2, s.get(2, s.get("2", 73)), s.get(1, s.get("1", 87))),
        (3, s.get(3, s.get("3", 59)), s.get(2, s.get("2", 73))),
        (4, s.get(4, s.get("4", 45)), s.get(3, s.get("3", 59))),
        (5, s.get(5, s.get("5", 20)), s.get(4, s.get("4", 45))),
    ]
    for grade, lower, upper in ranges:
        if pct >= lower:
            span = upper - lower
            if span <= 0:
                return float(grade)
            return round(grade + (upper - pct) / span, 1)
    return 6.0


def _build_student_pdf_single(student, questions, scan_map, session, config):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.lib.colors import HexColor

    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    pw, ph = A4

    weights = config.get("weights", {}) if config else {}
    scale_raw = config.get("grade_scale", DEFAULT_SCALE) if config else DEFAULT_SCALE
    scale = {int(k): v for k, v in scale_raw.items()}
    times = config.get("times", {}) if config else {}
    total_time = config.get("total_time") if config else None

    get_w = lambda qid: weights.get(str(qid), weights.get(qid, 1))
    max_score = sum(get_w(q["id"]) for q in questions if q["correct_answer"])

    score = 0
    results = []
    for q in questions:
        ans = scan_map.get((student["card_id"], q["id"]))
        correct = q["correct_answer"]
        is_correct = ans and correct and ans in correct
        w = get_w(q["id"])
        if is_correct:
            score += w
        results.append({"text": q["text"], "answer": ans, "correct": correct, "is_correct": is_correct, "weight": w})

    pct = round(score / max_score * 100) if max_score > 0 else 0
    grade = _decimal_grade(pct, scale)

    y = ph - 30 * mm
    c.setFont("Helvetica-Bold", 18)
    c.drawString(20 * mm, y, f"Auswertung — {student['name']}")
    y -= 8 * mm
    c.setFont("Helvetica", 11)
    c.drawString(20 * mm, y, session.name or f"Session #{session.id}")
    if total_time:
        c.drawRightString(pw - 20 * mm, y, f"Dauer: {total_time // 60}:{total_time % 60:02d}")
    y -= 12 * mm

    c.setFont("Helvetica-Bold", 14)
    c.drawString(20 * mm, y, f"Note: {grade}    —    {score}/{max_score} Punkte ({pct}%)")
    y -= 14 * mm

    # Table header
    c.setFont("Helvetica-Bold", 9)
    cols = [20 * mm, 35 * mm, pw - 80 * mm, pw - 55 * mm, pw - 35 * mm]
    c.drawString(cols[0], y, "#")
    c.drawString(cols[1], y, "Frage")
    c.drawString(cols[2], y, "Antwort")
    c.drawString(cols[3], y, "Lösung")
    c.drawString(cols[4], y, "Punkte")
    y -= 2 * mm
    c.setStrokeColorRGB(0, 0, 0)
    c.line(20 * mm, y, pw - 20 * mm, y)
    y -= 5 * mm

    c.setFont("Helvetica", 9)
    for i, r in enumerate(results):
        if y < 25 * mm:
            c.showPage()
            y = ph - 25 * mm
            c.setFont("Helvetica", 9)

        if r["is_correct"]:
            c.setFillColor(HexColor("#0a7d3e"))
        elif r["answer"] and r["correct"]:
            c.setFillColor(HexColor("#d1350f"))
        else:
            c.setFillColorRGB(0.4, 0.4, 0.4)

        text = strip_latex(r["text"])
        text = text[:60] + ("…" if len(text) > 60 else "")
        c.drawString(cols[0], y, str(i + 1))
        c.drawString(cols[1], y, text)
        c.drawString(cols[2], y, r["answer"] or "–")
        c.drawString(cols[3], y, r["correct"] or "–")
        pts = r["weight"] if r["is_correct"] else 0
        c.drawString(cols[4], y, str(pts))
        c.setFillColorRGB(0, 0, 0)
        y -= 5 * mm

    c.save()
    buf.seek(0)
    return buf


@router.get("/sessions/{session_id}/student-pdf/{card_id}")
async def student_evaluation_pdf(session_id: int, card_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    session = await db.get(Session, session_id)
    if not session:
        raise HTTPException(404)
    if session.owner_id and session.owner_id != user.id:
        raise HTTPException(403, "Kein Zugriff auf diese Session")

    student = None
    if session.class_id:
        result = await db.execute(
            select(Student).where(Student.class_id == session.class_id, Student.card_id == card_id)
        )
        s = result.scalar_one_or_none()
        if s:
            student = {"card_id": s.card_id, "name": s.name}
    if not student:
        raise HTTPException(404, "Lernende/r nicht gefunden")

    questions = []
    qmap = session.question_map or {}
    if session.question_set_id:
        result = await db.execute(
            select(QuestionSetItem).options(selectinload(QuestionSetItem.question))
            .where(QuestionSetItem.question_set_id == session.question_set_id)
            .order_by(QuestionSetItem.position)
        )
        for item in result.scalars().all():
            q = item.question
            questions.append({"id": q.id, "text": q.text, "correct_answer": qmap.get(str(q.id), q.correct_answer)})

    result = await db.execute(select(Scan).where(Scan.session_id == session_id))
    scan_map = {(s.student_id, s.question_id): s.answer for s in result.scalars().all()}

    config = session.eval_config or {}
    questions = [q for q in questions if (student["card_id"], q["id"]) in scan_map]
    buf = _build_student_pdf_single(student, questions, scan_map, session, config)
    filename = f"Auswertung_{student['name']}_{session_id}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/sessions/{session_id}/all-students-pdf")
async def all_students_pdf(session_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    session = await db.get(Session, session_id)
    if not session:
        raise HTTPException(404)
    if session.owner_id and session.owner_id != user.id:
        raise HTTPException(403, "Kein Zugriff auf diese Session")

    students = []
    if session.class_id:
        result = await db.execute(
            select(Student).where(Student.class_id == session.class_id).order_by(Student.name)
        )
        students = [{"card_id": s.card_id, "name": s.name} for s in result.scalars().all()]

    questions = []
    qmap = session.question_map or {}
    if session.question_set_id:
        result = await db.execute(
            select(QuestionSetItem).options(selectinload(QuestionSetItem.question))
            .where(QuestionSetItem.question_set_id == session.question_set_id)
            .order_by(QuestionSetItem.position)
        )
        for item in result.scalars().all():
            q = item.question
            questions.append({"id": q.id, "text": q.text, "correct_answer": qmap.get(str(q.id), q.correct_answer)})

    result = await db.execute(select(Scan).where(Scan.session_id == session_id))
    all_scans = result.scalars().all()
    scan_map = {(s.student_id, s.question_id): s.answer for s in all_scans}

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.lib.colors import HexColor

    config = session.eval_config or {}
    weights = config.get("weights", {})
    scale_raw = config.get("grade_scale", DEFAULT_SCALE)
    scale = {int(k): v for k, v in scale_raw.items()}
    get_w = lambda qid: weights.get(str(qid), weights.get(qid, 1))
    max_score = sum(get_w(q["id"]) for q in questions if q["correct_answer"])

    present = [s for s in students if any((s["card_id"], q["id"]) in scan_map for q in questions)]

    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    pw, ph = A4

    for si, student in enumerate(present):
        y = ph - 30 * mm
        c.setFont("Helvetica-Bold", 16)
        c.drawString(20 * mm, y, student["name"])
        y -= 7 * mm
        c.setFont("Helvetica", 10)
        c.drawString(20 * mm, y, session.name or f"Session #{session.id}")
        y -= 10 * mm

        student_questions = [q for q in questions if (student["card_id"], q["id"]) in scan_map]
        score = 0
        results = []
        for q in student_questions:
            ans = scan_map.get((student["card_id"], q["id"]))
            correct = q["correct_answer"]
            is_correct = ans and correct and ans in correct
            w = get_w(q["id"])
            if is_correct:
                score += w
            results.append({"text": q["text"], "answer": ans, "correct": correct, "is_correct": is_correct, "weight": w})

        student_max = sum(get_w(q["id"]) for q in student_questions if q["correct_answer"])
        pct = round(score / student_max * 100) if student_max > 0 else 0
        grade = _decimal_grade(pct, scale)

        c.setFont("Helvetica-Bold", 13)
        c.drawString(20 * mm, y, f"Note: {grade}    —    {score}/{student_max} Punkte ({pct}%)")
        y -= 12 * mm

        c.setFont("Helvetica-Bold", 9)
        c.drawString(20 * mm, y, "#")
        c.drawString(30 * mm, y, "Frage")
        c.drawString(pw - 65 * mm, y, "Antw.")
        c.drawString(pw - 45 * mm, y, "Lösung")
        c.drawString(pw - 28 * mm, y, "Pkt.")
        y -= 2 * mm
        c.line(20 * mm, y, pw - 20 * mm, y)
        y -= 5 * mm

        c.setFont("Helvetica", 9)
        for i, r in enumerate(results):
            if r["is_correct"]:
                c.setFillColor(HexColor("#0a7d3e"))
            elif r["answer"] and r["correct"]:
                c.setFillColor(HexColor("#d1350f"))
            else:
                c.setFillColorRGB(0.4, 0.4, 0.4)

            text = strip_latex(r["text"])
            text = text[:55] + ("…" if len(text) > 55 else "")
            c.drawString(20 * mm, y, str(i + 1))
            c.drawString(30 * mm, y, text)
            c.drawString(pw - 65 * mm, y, r["answer"] or "–")
            c.drawString(pw - 45 * mm, y, r["correct"] or "–")
            pts = r["weight"] if r["is_correct"] else 0
            c.drawString(pw - 28 * mm, y, str(pts))
            c.setFillColorRGB(0, 0, 0)
            y -= 5 * mm

        if si < len(present) - 1:
            c.showPage()

    c.save()
    buf.seek(0)
    filename = f"CardVote_Auswertungen_{session_id}.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})


@router.get("/classes/{class_id}/all-tests-student-pdf/{card_id}")
async def class_student_pdf(class_id: int, card_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    school_class = await db.get(SchoolClass, class_id)
    if not school_class:
        raise HTTPException(404)
    if school_class.owner_id and school_class.owner_id != user.id:
        raise HTTPException(403, "Kein Zugriff auf diese Klasse")

    result = await db.execute(select(Student).where(Student.class_id == class_id, Student.card_id == card_id))
    student_obj = result.scalar_one_or_none()
    if not student_obj:
        raise HTTPException(404, "Lernende/r nicht gefunden")

    result = await db.execute(select(Session).where(Session.class_id == class_id).order_by(Session.created_at))
    sessions = result.scalars().all()

    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as pdf_canvas
    from reportlab.lib.colors import HexColor

    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    pw, ph = A4

    y = ph - 30 * mm
    c.setFont("Helvetica-Bold", 18)
    c.drawString(20 * mm, y, f"Gesamtübersicht — {student_obj.name}")
    y -= 7 * mm
    c.setFont("Helvetica", 11)
    c.drawString(20 * mm, y, f"Klasse {school_class.name}")
    y -= 14 * mm

    # Table header
    c.setFont("Helvetica-Bold", 10)
    c.drawString(20 * mm, y, "Test")
    c.drawString(pw - 85 * mm, y, "Punkte")
    c.drawString(pw - 55 * mm, y, "%")
    c.drawString(pw - 35 * mm, y, "Note")
    y -= 2 * mm
    c.line(20 * mm, y, pw - 20 * mm, y)
    y -= 6 * mm

    total_score = 0
    total_max = 0
    test_count = 0

    c.setFont("Helvetica", 10)
    for session in sessions:
        questions = []
        qmap = session.question_map or {}
        if session.question_set_id:
            q_result = await db.execute(
                select(QuestionSetItem).options(selectinload(QuestionSetItem.question))
                .where(QuestionSetItem.question_set_id == session.question_set_id)
                .order_by(QuestionSetItem.position)
            )
            for item in q_result.scalars().all():
                q = item.question
                questions.append({"id": q.id, "correct_answer": qmap.get(str(q.id), q.correct_answer)})

        scan_result = await db.execute(select(Scan).where(Scan.session_id == session.id, Scan.student_id == card_id))
        scans = {s.question_id: s.answer for s in scan_result.scalars().all()}

        if not scans:
            continue

        config = session.eval_config or {}
        weights = config.get("weights", {})
        scale_raw = config.get("grade_scale", DEFAULT_SCALE)
        scale = {int(k): v for k, v in scale_raw.items()}
        get_w = lambda qid: weights.get(str(qid), weights.get(qid, 1))

        max_sc = sum(get_w(q["id"]) for q in questions if q["correct_answer"])
        sc = 0
        for q in questions:
            ans = scans.get(q["id"])
            if ans and q["correct_answer"] and ans in q["correct_answer"]:
                sc += get_w(q["id"])

        pct = round(sc / max_sc * 100) if max_sc > 0 else 0
        grade = _decimal_grade(pct, scale)

        total_score += sc
        total_max += max_sc
        test_count += 1

        # Get set name
        set_name = session.name
        if session.question_set_id:
            qs = await db.get(QuestionSet, session.question_set_id)
            if qs:
                set_name = qs.name

        if y < 25 * mm:
            c.showPage()
            y = ph - 25 * mm
            c.setFont("Helvetica", 10)

        label = set_name[:45] + ("…" if len(set_name) > 45 else "")
        c.drawString(20 * mm, y, label)
        c.drawString(pw - 85 * mm, y, f"{sc}/{max_sc}")
        c.drawString(pw - 55 * mm, y, f"{pct}%")

        g_color = "#0a7d3e" if grade <= 2 else "#b8860b" if grade <= 4 else "#d1350f"
        c.setFillColor(HexColor(g_color))
        c.drawString(pw - 35 * mm, y, str(grade))
        c.setFillColorRGB(0, 0, 0)
        y -= 6 * mm

    # Summary
    y -= 4 * mm
    c.line(20 * mm, y, pw - 20 * mm, y)
    y -= 8 * mm
    total_pct = round(total_score / total_max * 100) if total_max > 0 else 0
    total_grade = _decimal_grade(total_pct)
    c.setFont("Helvetica-Bold", 12)
    c.drawString(20 * mm, y, f"Gesamt: {total_score}/{total_max} ({total_pct}%)    Note: {total_grade}    ({test_count} Tests)")

    c.save()
    buf.seek(0)
    filename = f"CardVote_{student_obj.name}_Gesamtübersicht.pdf"
    return StreamingResponse(buf, media_type="application/pdf",
                             headers={"Content-Disposition": f'attachment; filename="{filename}"'})
